import math
from datetime import datetime
import os
from platform import python_version
import openpyxl
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill

start_time = datetime.now()

# Octant finder function
def octant_finder(u, v, w):
    '''
        Returns the octant using coordinates. We pass 3 coordinates to our function. The 1st and 2nd coordinate are used for finding the quadrant whereas the 3rd coordinate is used to determine the octant or sign from quadrant.  
    '''

    quadrant = 0

    if u >= 0:
        quadrant = 1 if v >= 0 else 4
    else:
        quadrant = 2 if v >= 0 else 3

    octant = quadrant
    if w < 0:
        octant = octant*(-1)

    return octant

# Data preprocessing function
def data_preprocessing(df):
    '''
        This function is used to format our dataframe and find the mean of U,V and W respectively for calculations. Also here we find the new values, U', V' and W', by subtracting the mean from original values and storing it. We find the values of octant for each coordinate using our 'octant_finder' function. 
    '''

    last = len(df)

    # Here we create new columns for mean of U,V and Z respectively
    df.insert(4, "U Avg", "", True)
    df.insert(5, "V Avg", "", True)
    df.insert(6, "W Avg", "", True)

    # We create a list to store average of U,V and W respectively
    avg = []

    # Here we find the mean
    for i in range(1, 4):
        t_avg = df.iloc[:, i].mean()
        t_avg = round(t_avg,3)
        df.iloc[0, i+3] = t_avg
        avg.append(df.iloc[0, i+3])


    # We create our new columns for U',V' and Z' using U,V and Z, and their means respectively
    df.insert(7, "U'=U - U avg", round(df.iloc[:, 1]-avg[0],3), True)
    df.insert(8, "V'=V - V avg", round(df.iloc[:, 2]-avg[1],3), True)
    df.insert(9, "W'=W - W avg", round(df.iloc[:, 3]-avg[2],3), True)

    # Here we create a new column to find the
    df.insert(10, "Octant", 0, True)
    # Now we use the apply to pass a function and apply it on every value of our "Octant" column
    df["Octant"] = df.apply(lambda row: octant_finder(
        row["U'=U - U avg"], row["V'=V - V avg"], row["W'=W - W avg"]), axis=1)

    df.insert(11, "", "", True)

    return

# Count table maker function
def count_table_maker(df, mod):
    '''
        This function is used to format our dataframe so that we can insert the count of octants for the given mod.
    '''

    df.insert(12, "", "", True)
    df.iloc[2, 12] = "Mod "+str(mod)

    last = len(df)

    df.insert(13, "Overall Octant Count", "", True)
    df.iloc[1, 13] = "Octant ID"
    df.iloc[2, 13] = "Overall Count"

    # We find the range using the given mod
    n = int(last/mod)
    if (last % mod):
        n += 1

    for i in range(0, n):
        l = (mod*i)
        r = min(mod*(i+1)-1, last-1)
        df.loc[i+3, "Overall Octant Count"] = str(l)+"-"+str(r)

    # We create columns for each octant value so that we can store its count
    # start = 14
    for i in range(14, 22):
        y = (int)((i/2)-6)
        if(i % 2):
            y *= -1
        df.insert(i, "", "", True)
        df.iloc[1, i] = y

    return

# Octant counter
def octant_counter(df, mod):
    '''
        This function is used to store the count of octants for the a given mod and also for the whole dataset.
    '''

    # We create a series which stores the count of each individual value using 'value_counts'
    count_all = df.iloc[:, 10].value_counts()

    last = len(df)

    # We find the range using the given mod
    n = int(last/mod)
    if (last % mod):
        n += 1

    # We iterate for each octant value
    for i in range(-4, 5):

        # If we have any value of j other than '0' that means we can use it for our octant count
        if(i):
            col = 2*(abs(i)-1)
            if(i < 0):
                col += 1

            # Storing the count of each octant in that mod group
            if(i in count_all):
                df.iloc[2, 14+col] = count_all[i]
            else:
                df.iloc[2, 14+col] = 0

    # We iterate for each mod group
    for i in range(0, n):
        # We create a series which stores the count of each individual value using 'value_counts'
        count_mod = df.iloc[(i*mod):((i+1)*mod), 10].value_counts()

        # We iterate for each octant value
        for j in range(-4, 5):

            # If we have any value of j other than '0' that means we can use it for our octant count
            if(j):
                col = 2*(abs(j)-1)
                if(j < 0):
                    col += 1

                # Storing the count of each octant in that mod group
                if(j in count_mod):
                    df.iloc[i+3, 14+col] = count_mod[j]
                else:
                    df.iloc[i+3, 14+col] = 0

    return

# Rank table filler function
def rank_filler(df, octant_name, ind, n, cell_highlight):

    octants = [1, -1, 2, -2, 3, -3, 4, -4]

    store = []

    cells = ["W","X","Y","Z","AA","AB","AC","AD"]

    for i in range(1, 9):
        val = df.iloc[ind, 13+i]
        store.append((val, i))

    store.sort()
    count = 1
    for i in store:
        df.iloc[ind, 21+i[1]] = 9-count
        if(count == 8):
            df.iloc[ind, 30] = octants[i[1]-1]
            df.iloc[ind, 31] = octant_name[i[1]-1]
            col_name = cells[i[1]-1] + str(ind+2)
            cell_highlight.append(col_name)

        count += 1

    if(ind != 2):
        df.iloc[6+n+store[7][1], 30] += 1

    return

# Rank table maker
def rank_table_maker(df, mod, octant_name, cell_highlight):

    last = len(df)

    # We find the range using the given mod
    n = int(last/mod)
    if (last % mod):
        n += 1

    #
    for i in range(1, 9):

        head = math.ceil((i/2))
        if(i % 2 == 0):
            head *= -1

        df.insert(21+i, "", "", True)
        df.iloc[1, 21+i] = "Rank "+str(head)

    df.insert(30, "", "", True)
    df.iloc[1, 30] = "Rank1 Octant ID"
    df.insert(31, "", "", True)
    df.iloc[1, 31] = "Rank1 Octant Name"

    df.iloc[6+n, 28] = "Octant ID"
    df.iloc[6+n, 29] = "Octant Name"
    df.iloc[6+n, 30] = "Count of Rank 1 Mod Values"

    for i in range(1, 9):

        head = math.ceil((i/2))
        if(i % 2 == 0):
            head *= -1

        df.iloc[6+n+i, 28] = head
        df.iloc[6+n+i, 29] = octant_name[i-1]
        df.iloc[6+n+i, 30] = 0

    # # Overall rank calculation
    rank_filler(df, octant_name, 2, n,cell_highlight)

    # # Rank calculation for each octant range
    for i in range(0, n):
        rank_filler(df, octant_name, 3+i, n,cell_highlight)

    return

# Rank tables
def rank_function(df, mod, cell_highlight):

    data_preprocessing(df)
    count_table_maker(df, mod)
    octant_counter(df, mod)

    octant_name = [
        "Internal outward interaction",
        "External outward interaction",
        "External Ejection",
        "Internal Ejection",
        "External inward interaction",
        "Internal inward interaction",
        "Internal Sweep",
        "External Sweep"
    ]

    rank_table_maker(df, mod, octant_name,cell_highlight)


""" ------------------------------------------------------------------------------------------------------------------------------------------------------ """

# Table maker
def transition_table_maker(df, l, r, ind):
    """
        This function is used to make the desired table based on 'l' and 'r' provided.
    """

    # try:
    #     heading = "Mod Transition Count"

    #     # # If l < 0 then we are making the overall table else it is for individual ranges.
    #     # if l < 0:
    #     #     heading += "Overall "
    #     # else:
    #     #     heading += "Mod "
    #     #     df.iloc[ind+1,12] = str(l)+"-"+str(r)

    #     # heading += "Transition Count"
    #     if l >= 0:
    #         df.iloc[ind, 34] = heading
    #         df.iloc[ind+1, 34] = str(l)+"-"+str(r)

    #     if l < 0:
    #         ind -= 1

    #     df.iloc[ind+1, 35] = "To"
    #     # ind += 2
    #     df.iloc[ind+2, 34] = "Ocatant #"
    #     df.iloc[ind+3, 33] = "From"

    #     for i in range(-4, 5):
    #         if(i):
    #             x = 2*(abs(i)-1)
    #             if(i < 0):
    #                 x += 1
    #             df.iloc[ind+3+x, 34] = i
    #             df.iloc[ind+2, 35+x] = i

    #     # ind += 1

    #     # We set all the cells in the table to 0.
    #     for i in range(0, 8):
    #         for j in range(0, 8):
    #             df.iloc[ind+i+3, 35+j] = 0

    # except:
    #     print("There was an error while performing 'transition_table_maker()'.")
    #     exit()

    heading = "Mod Transition Count"

        # # If l < 0 then we are making the overall table else it is for individual ranges.
        # if l < 0:
        #     heading += "Overall "
        # else:
        #     heading += "Mod "
        #     df.iloc[ind+1,12] = str(l)+"-"+str(r)

        # heading += "Transition Count"
    if l >= 0:
        df.iloc[ind, 34] = heading
        df.iloc[ind+1, 34] = str(l)+"-"+str(r)

    if l < 0:
        ind -= 1

    df.iloc[ind+1, 35] = "To"
    # ind += 2
    df.iloc[ind+2, 34] = "Ocatant #"
    df.iloc[ind+3, 33] = "From"

    for i in range(-4, 5):
        if(i):
            x = 2*(abs(i)-1)
            if(i < 0):
                x += 1
            df.iloc[ind+3+x, 34] = i
            df.iloc[ind+2, 35+x] = i

    # ind += 1

    # We set all the cells in the table to 0.
    for i in range(0, 8):
        for j in range(0, 8):
            df.iloc[ind+i+3, 35+j] = 0

# Cell incrementer
def transition_values_finder(df, mod, l, r, ind, total):
    """
        This function is used to increment the cell for each 'from' and 'to' pair.
    """

    try:
        if(r != total):
            r += 1

        n = total//mod
        if(total % mod):
            n += 1

        for i in range(l, r):
            val1 = df.iloc[i, 10]
            val2 = df.iloc[i+1, 10]
            # Find the index of each column and row based on from and to octant value.
            col = 2*(abs(val2)-1)+35
            if(val2 < 0):
                col += 1
            row = 2*(abs(val1)-1)
            if(val1 < 0):
                row += 1

            # Incrementing the value in the overall and individual cell.
            df.iloc[row+2, col] += 1
            df.iloc[row+ind+3, col] += 1

    except:
        print("There was an error while performing 'transition_values_finder()'.")
        exit()

# Transition table filler
def transition_table_filler(df, mod, total, cell_highlight, transition_border):
    """
        We use this function to create and fill every individual table.
    """

    cells = ["AJ","AK","AL","AM","AN","AO","AP","AQ"]

    # try:
    #     n = total//mod
    #     if(total % mod):
    #         n += 1

    #     ind = 13

    #     for i in range(0, n):
    #         l = i*mod
    #         r = min(((i+1)*mod)-1, total)
    #         transition_table_maker(df, l, r, ind)
    #         transition_values_finder(df, mod, l, r, ind, total)

    #         transition_border[0] = ind+1

    #         ind_h = ind+3
    #         for j in  range(0,8):
    #             # [max,col]
    #             vals = [-1,-1]

    #             for k in range(0,8):
    #                 if df.iloc[ind_h+j,35+k] > vals[0]:
    #                     vals[0] = df.iloc[ind_h+j,35+k]
    #                     vals[1] = k

    #             cell = cells[vals[1]]+str(ind_h+j+2)
    #             cell_highlight.append(cell)

    #         ind += 13


    #     # Overall highlight

    #     for j in range(2,10):

    #         vals = [-1,-1]
    #         for k in range(0,8):
    #             if df.iloc[j,35+k] > vals[0]:
    #                 vals[0] = df.iloc[j,35+k]
    #                 vals[1] = k

    #         cell = cells[vals[1]]+str(j+2)
    #         cell_highlight.append(cell)



    # except:
    #     print("There was an error while performing 'transition_table_filler()'.")
    #     exit()

    n = total//mod
    if(total % mod):
        n += 1

    ind = 13

    for i in range(0, n):
        l = i*mod
        r = min(((i+1)*mod)-1, total)
        transition_table_maker(df, l, r, ind)
        transition_values_finder(df, mod, l, r, ind, total)

        transition_border[0] = ind+1

        ind_h = ind+3
        for j in  range(0,8):
            # [max,col]
            vals = [-1,-1]

            for k in range(0,8):
                if df.iloc[ind_h+j,35+k] > vals[0]:
                    vals[0] = df.iloc[ind_h+j,35+k]
                    vals[1] = k

            cell = cells[vals[1]]+str(ind_h+j+2)
            cell_highlight.append(cell)

        ind += 13


    # Overall highlight

    for j in range(2,10):

        vals = [-1,-1]
        for k in range(0,8):
            if df.iloc[j,35+k] > vals[0]:
                vals[0] = df.iloc[j,35+k]
                vals[1] = k

        cell = cells[vals[1]]+str(j+2)
        cell_highlight.append(cell)

# Overall function for transition table
def transition_function(df, mod, cell_highlight, transition_border):

    total = len(df)-1

    n = total//mod
    if(total % mod):
        n += 1

    df.insert(32, "", "", True)
    df.insert(33, "", "", True)
    df.insert(34, "Overall Transition Count", "", True)

    for i in range(0, 8):
        df.insert(35+i, "", "", True)

    transition_table_maker(df, -1, -1, 0)
    transition_table_filler(df, mod, total, cell_highlight,transition_border)


""" ------------------------------------------------------------------------------------------------------------------------------------------------------ """

#
def ls_table_maker(df):

    df.insert(43, "", "", True)
    df.insert(44, "Longest Subsequence Length", "", True)
    df.insert(45, "", "", True)
    df.insert(46, "", "", True)

    df.iloc[1, 44] = "Octant ##"
    df.iloc[1, 45] = "Longest Subsequence Length"
    df.iloc[1, 46] = "Count"

    for i in range(-4, 5):
        if(i):
            ind = 2*(abs(i)-1)
            if(i < 0):
                ind += 1
            df.iloc[ind+2, 44] = i
            df.iloc[ind+2, 45] = 0
            df.iloc[ind+2, 46] = 0

#
def ls_finder(df, store):

    l = len(df)-1

    currl = 1
    start = 0
    for i in range(1, l):

        # If current and next octant are same then we increase the length.
        if(df.iloc[i, 10] == df.iloc[i+1, 10]):
            currl += 1

        # Else we check if our length is longest or not and update the count and length of longest subsequence.
        else:
            num = df.iloc[i, 10]
            row = 2*(abs(num)-1)
            end = df.iloc[i, 0]
            if(num < 0):
                row += 1
            if(currl == df.iloc[row+2, 45]):
                df.iloc[row+2, 46] += 1
                store[num].append([start, end])
            elif (currl > df.iloc[row+2, 45]):
                df.iloc[row+2, 46] = 1
                df.iloc[row+2, 45] = currl
                store[num].clear()
                store[num].append([start, end])

            start = df.iloc[i+1, 0]
            currl = 1

    num = df.iloc[i, 10]
    row = 2*(abs(num)-1)
    if(num < 0):
        row += 1
        if(currl == df.iloc[row+2, 45]):
            df.iloc[row+2, 46] += 1
            store[num].append([start, end])
        elif (currl > df.iloc[row+2, 45]):
            df.iloc[row+2, 46] = 1
            df.iloc[row+2, 45] = currl
            store[num].clear()
            store[num].append([start, end])

#
def time_function(df, store, time_border):

    df.insert(47, "", "", True)
    df.insert(48, "Longest Subsequence Length with Range", "", True)
    df.insert(49, "", "", True)
    df.insert(50, "", "", True)

    df.iloc[1, 48] = "Octant ##"
    df.iloc[1, 49] = "Longest Subsequence Length"
    df.iloc[1, 50] = "Count"

    ind = 2

    for i in range(0, 8):
        num = ((i//2)+1)
        if(i % 2):
            num *= -1

        # Num
        df.iloc[ind, 48] = num
        longest_subsequence_length = df.iloc[i+2, 45]
        count = df.iloc[i+2, 46]
        df.iloc[ind, 49] = longest_subsequence_length
        df.iloc[ind, 50] = count
        ind += 1
        df.iloc[ind, 48] = "Time"
        df.iloc[ind, 49] = "From"
        df.iloc[ind, 50] = "To"
        ind += 1

        # For filling the time of longest subsequence
        for j in range(0, count):
            time_from = store[num][j][0]
            time_to = store[num][j][1]
            df.iloc[ind, 49] = time_from
            df.iloc[ind, 50] = time_to
            ind += 1

    time_border[0] = ind

#
def longest_subsequence_function(df, time_border):

    store = {
        1: [], -1: [], 2: [], -2: [], 3: [], -3: [], 4: [], -4: []
    }

    ls_table_maker(df)
    ls_finder(df, store)
    time_function(df, store, time_border)


""" ------------------------------------------------------------------------------------------------------------------------------------------------------ """

#
def main_border(wb, mod, last, time_border, cell_highlight, transition_border):

    ws = wb['Sheet1']

    all = Side(border_style='thin')
    border = Border(left=all, top=all, right=all, bottom=all)
    highlight = PatternFill(patternType='solid',fgColor='FFFF00')

    cols_rank = ["N", "O", "P", "Q", "R", "S", 'T', 'U', 'V',
                 "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF"]
    cols_oct = ["AC", "AD", "AE"]
    cols_lsl = ["AS", "AT", "AU"]
    cols_time = ["AW", "AX", "AY"]
    cols_transition = ["AJ","AK","AL","AM","AN","AO","AP","AQ"]

    n = int(last/mod)
    if (last % mod):
        n += 1

    for col in cols_rank:
        for i in range(3, 5+n):
            cell = col+str(i)
            ws[cell].border = border

    for col in cols_oct:
        for i in range(12, 21):
            cell = col+str(i)
            ws[cell].border = border

    for col in cols_lsl:
        for i in range(3, 12):
            cell = col+str(i)
            ws[cell].border = border

    for col in cols_time:
        for i in range(3, time_border[0]+2):
            cell = col+str(i)
            ws[cell].border = border

    for row in range(18,transition_border[0]+13,13):
        for i in range(0,8):
            for col in cols_transition:
                cell = col+str(row+i)
                ws[cell].border = border

    for row in range(4,12):
        for col in cols_transition:
            cell = col+str(row)
            ws[cell].border = border


    for cell in cell_highlight:
        ws[cell].fill = highlight
    

""" ------------------------------------------------------------------------------------------------------------------------------------------------------ """

# Help
def octant_analysis(name,mod=5000):
    
    directory = "input"
    # suffix = "_octant_analysis_mod_"+str(mod)+".xlsx"
    time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    suffix = "_mod_"+str(mod)+"_"+time+".xlsx"
    prefix = "output\\"
    # for filename in os.listdir(directory):
    #     f = os.path.join(directory, filename)
        
    #     name = f[6:-5]

    #     df = pd.read_excel(f)
    #     last = len(df)

    #     cell_highlight = []

    #     transition_border = [-1]

    #     rank_function(df, mod,cell_highlight)

    #     transition_function(df, mod, cell_highlight,transition_border)

    #     time_border = [-1]
        
    #     longest_subsequence_function(df, time_border)

    #     df.to_excel(prefix+name+suffix, index=False)

    #     wb = openpyxl.load_workbook(prefix+name+suffix)

    #     main_border(wb, mod, last, time_border,cell_highlight,transition_border)

    #     wb.save(prefix+name+suffix)
        
    df = pd.read_excel(name)
    last = len(df)

    cell_highlight = []

    transition_border = [-1]

    rank_function(df, mod,cell_highlight)
    
    transition_function(df, mod, cell_highlight,transition_border)

    time_border = [-1]
    
    longest_subsequence_function(df, time_border)

    if(type(name) != str):
        name = name.name
    else:
        t_name = os.path.basename(name)
        t_name = os.path.splitext(t_name)
        name = t_name[0]+t_name[1]

   
    file_name = prefix+name[:-5]+suffix
    
    df.to_excel(file_name, index=False)

    wb = openpyxl.load_workbook(file_name)
    
    main_border(wb, mod, last, time_border,cell_highlight,transition_border)

    wb.save(file_name)

    return file_name

# Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
# Save all the excel files in a the output/ folder. Only xlsx to be allowed
# output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename.

# Code
# ver = python_version()

# if ver == "3.8.10":
#     print("Correct Version Installed")
# else:
#     print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


# mod = 5000
# octant_analysis(mod)


# # This shall be the last lines of the code.
# end_time = datetime.now()
# print('Duration of Program Execution: {}'.format(end_time - start_time))
